# Скрипт выводит на терминал список всех снэпшотов в среде oVirt, а также производит экспорт списка в файл snapshots_data_YYYYMMDD_hhmmss.xlsx
# В выводе следующие поля:
# VM name - имя ВМ с которой сделан снэпшот
    # Snap name - имя снэпшота
    # Status - состояние снэпшота
    # ID - ID снэпшота
    # Disk - Имя диска снэпшота
    # Storage - Домен хранения дисков снэпшота

# Перед началом работы со скриптом задать следующие переменные:
    # OVIRT_USER - Пользователь Ovirt
    # OVIRT_PASS - Пароль пользователя Ovirt
    # OVIRT_URL - URL локального ovirt-engine

import ovirtsdk4 as sdk
import ovirtsdk4.types as types
import time, requests
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

OVIRT_USER = "admin@internal"  # admin
OVIRT_PASS = "password"
OVIRT_URL = "https://engine.redvirt.tst/ovirt-engine/api"

def ovirt_connect(OVIRT_URL):
    connection = sdk.Connection(
        url=OVIRT_URL,
        username=OVIRT_USER,
        password=OVIRT_PASS,
        insecure=True,
    )
    return connection

# Функция поиска снэпшотов
def find_snap(connection):
    # Получаем ссылку на корневую службу:
    system_service = connection.system_service()

    # Находим все ВМ и сохраняем их ID и названия в словаре :
    vms_service = system_service.vms_service()
    vms_map = {
        vm.id: vm.name
        for vm in vms_service.list()
    }

    # Находим все сторадж домены и сохраняем их ID и названия в словаре :
    sds_service = system_service.storage_domains_service()
    sds_map = {
        sd.id: sd.name
        for sd in sds_service.list()
    }

    print("{:<20} | {:<20} | {:<20} | {:<20} | {:<20} | {:<20}".format('VM name', 'Snap name', 'Status', 'ID', 'Disk', 'Storage'))

    # Задаем пустой список, в который далее будем писать данные для экспорта в эксель
    data = []

    # Для каждой ВМ ищем ее снэпшот , затем для каждого снэпшота ищем его диски:
    for vm_id, vm_name in vms_map.items():
        vm_service = vms_service.vm_service(vm_id)
        snaps_service = vm_service.snapshots_service()
        snaps_map = {
            snap.id: snap.description
            for snap in snaps_service.list()
        }
        for snap_id, snap_description in snaps_map.items():
            snap_service = snaps_service.snapshot_service(snap_id)
            disks_service = snap_service.disks_service()
            status = snap_service.get()

            for disk in disks_service.list():
                if len(disk.storage_domains) > 0:
                    sd_id = disk.storage_domains[0].id
                    sd_name = sds_map[sd_id]
                    print("{:<20} | {:<20} | {:<20} | {:<20} | {:<20} | {:<20}".format(
                        vm_name, snap_description, status.snapshot_status, snap_id, disk.alias, sd_name
                    ))
                    data.append({
                        "VM name": vm_name,
                        "Snap name": snap_description,
                        "Status": str(status.snapshot_status),  
                        "ID": snap_id,
                        "Disk": disk.alias,
                        "Storage": sd_name,
                    })

    # Эскпорт в эксель с помошью openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Snapshots Data"

    # Пишем заголовок в переменную
    headers = ["VM name", "Snap name", "Status", "ID", "Disk", "Storage"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    # Пишем данные в переменную
    for row_num, row_data in enumerate(data, 2):
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            value = row_data[header]
            ws[f"{col_letter}{row_num}"] = str(value)  

    # Добавляем  метку времени в экспортируемый файл
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"snapshots_data_{timestamp}.xlsx"
    wb.save(filename)
    print(f"Данные сохранены в файл {filename}")

def main():
    connection = ovirt_connect(OVIRT_URL)
    find_snap(connection)
    connection.close()

if __name__ == '__main__':
    main()